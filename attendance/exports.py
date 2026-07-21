from openpyxl import Workbook
from reportlab.platypus import SimpleDocTemplate
from reportlab.platypus import Table
from reportlab.platypus import TableStyle
from reportlab.lib import colors

from django.http import HttpResponse

from .models import Attendance


def export_excel(request):

    workbook = Workbook()

    sheet = workbook.active

    sheet.title = "Attendance"

    headers = [

        "Employee ID",

        "Employee Name",

        "Date",

        "Check In",

        "Check Out",

        "Status"

    ]

    for column, header in enumerate(headers, 1):

        sheet.cell(
            row=1,
            column=column
        ).value = header

    attendance = Attendance.objects.select_related(
        "employee"
    )

    row = 2

    for record in attendance:

        sheet.cell(
            row=row,
            column=1
        ).value = record.employee.employee_id

        sheet.cell(
            row=row,
            column=2
        ).value = f"{record.employee.first_name} {record.employee.last_name}"

        sheet.cell(
            row=row,
            column=3
        ).value = str(record.date)

        sheet.cell(
            row=row,
            column=4
        ).value = str(record.check_in)

        sheet.cell(
            row=row,
            column=5
        ).value = str(record.check_out)

        sheet.cell(
            row=row,
            column=6
        ).value = record.status

        row += 1

    response = HttpResponse(

        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"

    )

    response["Content-Disposition"] = 'attachment; filename="Attendance.xlsx"'

    workbook.save(response)

    return response


def export_pdf(request):

    response = HttpResponse(

        content_type="application/pdf"

    )

    response["Content-Disposition"] = 'attachment; filename="Attendance.pdf"'

    document = SimpleDocTemplate(
        response
    )

    attendance = Attendance.objects.select_related(
        "employee"
    )

    data = [[

        "Employee",

        "Date",

        "Check In",

        "Check Out",

        "Status"

    ]]

    for record in attendance:

        data.append([

            record.employee.employee_id,

            str(record.date),

            str(record.check_in),

            str(record.check_out),

            record.status

        ])

    table = Table(data)

    table.setStyle(

        TableStyle([

            ("BACKGROUND", (0,0), (-1,0), colors.grey),

            ("TEXTCOLOR", (0,0), (-1,0), colors.white),

            ("GRID", (0,0), (-1,-1), 1, colors.black),

            ("BACKGROUND", (0,1), (-1,-1), colors.beige),

            ("ALIGN", (0,0), (-1,-1), "CENTER"),

        ])

    )

    document.build([table])

    return response